#!/usr/bin/env python3
"""Backup workbook from the shared book. Never writes the 17 Aug original."""
from __future__ import annotations

import json
import os
import shutil
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
ORIG = Path.home() / "Developer/Forge-Capital/260817 Master Investor Tracker TF (CANONICAL).xlsx"
ENV = ROOT / ".env.local"

def load_env() -> dict[str, str]:
    env: dict[str, str] = {}
    if not ENV.exists():
        return env
    for line in ENV.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def rest_all(url: str, key: str, schema: str, table: str, select: str) -> list:
    out = []
    start = 0
    while True:
        req = urllib.request.Request(
            f"{url}/rest/v1/{table}?select={select}&offset={start}&limit=1000"
        )
        req.add_header("apikey", key)
        req.add_header("Authorization", f"Bearer {key}")
        req.add_header("Accept-Profile", schema)
        req.add_header("Range", f"{start}-{start+999}")
        with urllib.request.urlopen(req, timeout=60) as resp:
            batch = json.loads(resp.read().decode())
        out.extend(batch)
        if len(batch) < 1000:
            break
        start += 1000
    return out


env = load_env()
url, key = env.get("FORGE_CAPITAL_DB_URL"), env.get("FORGE_CAPITAL_DB_SERVICE_ROLE")
if not url or not key:
    raise SystemExit("FORGE_CAPITAL_DB_* missing")
if not ORIG.exists():
    raise SystemExit(f"template missing {ORIG}")

stamp_before = ORIG.stat().st_mtime
firms = rest_all(url, key, "core", "firms", "id,canonical_name,website_domain,sectors,dnc,notes")
people = rest_all(
    url, key, "core", "people",
    "id,firm_id,full_name,email,email_state,dnc,linkedin_url,role_title,provenance",
)
parts = rest_all(
    url, key, "engage", "participations",
    "firm_id,person_id,mandate_id,stage,status_note,first_sent,latest_touch",
)
mandates = rest_all(url, key, "engage", "mandates", "id,code")
code_by = {m["id"]: m["code"] for m in mandates}
people_by: dict[str, list] = {}
for p in people:
    people_by.setdefault(p["firm_id"], []).append(p)
parts_by: dict[str, list] = {}
for p in parts:
    parts_by.setdefault(p["firm_id"], []).append(p)

yymmdd = datetime.now(timezone.utc).strftime("%y%m%d")
out = Path.home() / f"Developer/Forge-Capital/{yymmdd} Master Investor Tracker TF (CANONICAL).xlsx"
if out.resolve() == ORIG.resolve() or "260817" in out.name:
    raise SystemExit("refusing to write the 17 Aug original")
shutil.copy2(ORIG, out)

wb = openpyxl.load_workbook(out)
ws = wb["Master Tracker"]
# Keep header rows 1-2. Clear old data from row 3 down.
if ws.max_row > 2:
    ws.delete_rows(3, ws.max_row - 2)

CODES = ["SK", "FF", "PA", "SS", "CA", "US", "OD", "HO"]
# 1-based Excel columns: Latest in each block
LATEST = [17, 22, 27, 32, 37, 42, 47, 52]
DAYS = [18, 23, 28, 33, 38, 43, 48, 53]

sorted_firms = sorted(firms, key=lambda f: (f.get("canonical_name") or "").lower())
for i, f in enumerate(sorted_firms):
    r = i + 3
    plist = people_by.get(f["id"]) or []
    named = next((p for p in plist if not p.get("dnc")), plist[0] if plist else None)
    pp = parts_by.get(f["id"]) or []
    by_code = {}
    for p in pp:
        c = code_by.get(p["mandate_id"])
        if c:
            by_code[c] = p
    ws.cell(r, 10, f.get("canonical_name"))
    ws.cell(r, 11, f.get("website_domain"))
    ws.cell(r, 12, named.get("full_name") if named else None)
    ws.cell(r, 13, named.get("email") if named else None)
    sectors = f.get("sectors")
    sector = ", ".join(sectors) if isinstance(sectors, list) else sectors
    if f.get("dnc"):
        sector = " · ".join(x for x in [sector, "DNC"] if x)
    ws.cell(r, 14, sector)
    ws.cell(r, 15, len(pp) or None)
    max_bits = []
    for j, code in enumerate(CODES):
        p = by_code.get(code)
        tick_col = 2 + j
        first_col = 16 + j * 5
        latest_col = first_col + 1
        days_col = first_col + 2
        status_col = first_col + 3
        comm_col = first_col + 4
        latest_letter = openpyxl.utils.get_column_letter(latest_col)
        max_bits.append(f"{latest_letter}{r}")
        if not p:
            ws.cell(r, days_col, f'=IF({latest_letter}{r}="","",TODAY()-{latest_letter}{r})')
            continue
        ws.cell(r, tick_col, "✓")
        first = (p.get("first_sent") or "")[:10] or None
        latest = (p.get("latest_touch") or "")[:10] or None
        ws.cell(r, first_col, first)
        ws.cell(r, latest_col, latest)
        ws.cell(r, days_col, f'=IF({latest_letter}{r}="","",TODAY()-{latest_letter}{r})')
        ws.cell(r, status_col, p.get("status_note") or p.get("stage"))
        ws.cell(r, comm_col, p.get("status_note"))
    mx = ",".join(max_bits)
    ws.cell(r, 1, f'=IF(MAX({mx})=0,"",TODAY()-MAX({mx}))')

def clear_sheet(ws, header_rows: int) -> None:
    if ws.max_row > header_rows:
        ws.delete_rows(header_rows + 1, ws.max_row - header_rows)


def write_rows(ws, start_row: int, rows: list[list]) -> None:
    for i, row in enumerate(rows):
        for c, val in enumerate(row, 1):
            ws.cell(start_row + i, c, val)


firms_by = {f["id"]: f for f in firms}
people_by_id = {p["id"]: p for p in people}
od_id = next((m["id"] for m in mandates if m.get("code") == "OD"), None)
us_id = next((m["id"] for m in mandates if m.get("code") == "US"), None)
od_parts = [p for p in parts if p.get("mandate_id") == od_id]
us_parts = [p for p in parts if p.get("mandate_id") == us_id]


def firm_name(fid):
    f = firms_by.get(fid) or {}
    return f.get("canonical_name")


def firm_web(fid):
    f = firms_by.get(fid) or {}
    return f.get("website_domain")


def person_name(pid):
    p = people_by_id.get(pid) or {}
    return p.get("full_name")


def person_email(pid):
    p = people_by_id.get(pid) or {}
    return p.get("email")


def od_row(p):
    return [
        firm_name(p.get("firm_id")),
        firm_web(p.get("firm_id")),
        p.get("stage"),
        (p.get("latest_touch") or p.get("first_sent") or "")[:10] or None,
        p.get("status_note"),
        person_name(p.get("person_id")),
        person_email(p.get("person_id")),
        p.get("status_note"),
    ]


if "Dashboard" in wb.sheetnames:
    d = wb["Dashboard"]
    d["A1"] = "Shared-book backup — generated from Corpus"
    d["A2"] = datetime.now(timezone.utc).isoformat()
    d["A3"] = f"Firms {len(firms)} · people {len(people)} · participations {len(parts)}"
    d["A4"] = "The 17 Aug original was not modified."

if "README" in wb.sheetnames:
    r = wb["README"]
    r["A1"] = "Generated from the shared book. Excel is a download only."
    r["A2"] = "Do not type in the 17 Aug CANONICAL file."

if "US Exploration (Forge)" in wb.sheetnames:
    us = wb["US Exploration (Forge)"]
    clear_sheet(us, 2)
    write_rows(
        us,
        3,
        [
            [
                firm_name(p.get("firm_id")),
                person_name(p.get("person_id")),
                person_email(p.get("person_id")),
                "book",
                p.get("stage"),
                "",
                p.get("status_note"),
            ]
            for p in us_parts
        ],
    )

if "Odysseus — Jordan full" in wb.sheetnames:
    j = wb["Odysseus — Jordan full"]
    clear_sheet(j, 1)
    write_rows(
        j,
        2,
        [
            [
                firm_name(p.get("firm_id")),
                firm_web(p.get("firm_id")),
                "",
                p.get("stage"),
                p.get("status_note"),
                firm_name(p.get("firm_id")),
                "book",
                person_name(p.get("person_id")),
                person_email(p.get("person_id")),
                "",
                "",
                p.get("status_note"),
            ]
            for p in od_parts
        ],
    )

do_not = [
    p for p in od_parts
    if (p.get("stage") in ("disqualified", "blocked", "closed_lost"))
    or "do not" in (p.get("status_note") or "").lower()
    or "cautious" in (p.get("status_note") or "").lower()
    and "leave" in (p.get("status_note") or "").lower()
]
low = [p for p in od_parts if "cautious" in (p.get("status_note") or "").lower()]
cands = [p for p in od_parts if (p.get("stage") or "") in ("research", "awaiting_signoff")]

if "Odysseus — DO NOT outreach" in wb.sheetnames:
    ws = wb["Odysseus — DO NOT outreach"]
    clear_sheet(ws, 1)
    write_rows(ws, 2, [od_row(p) for p in do_not])

if "Odysseus — Low priority OK" in wb.sheetnames:
    ws = wb["Odysseus — Low priority OK"]
    clear_sheet(ws, 1)
    write_rows(ws, 2, [od_row(p) for p in low])

if "Odysseus — Candidates approval" in wb.sheetnames:
    ws = wb["Odysseus — Candidates approval"]
    clear_sheet(ws, 1)
    write_rows(
        ws,
        2,
        [
            [
                firm_name(p.get("firm_id")),
                firm_web(p.get("firm_id")),
                person_name(p.get("person_id")),
                "",
                "",
                "",
                "",
                "",
                "",
                p.get("status_note"),
            ]
            for p in cands
        ],
    )

if "Odysseus — Jordan DO-NOT-APPROA" in wb.sheetnames:
    ws = wb["Odysseus — Jordan DO-NOT-APPROA"]
    clear_sheet(ws, 1)
    write_rows(
        ws,
        2,
        [
            [
                firm_name(p.get("firm_id")),
                firm_web(p.get("firm_id")),
                person_name(p.get("person_id")),
                person_email(p.get("person_id")),
                p.get("stage"),
                firm_web(p.get("firm_id")),
            ]
            for p in do_not
        ],
    )

li = [p for p in people if (p.get("provenance") or "").lower() == "linkedin"]
if "LinkedIn Connections" in wb.sheetnames:
    ws = wb["LinkedIn Connections"]
    clear_sheet(ws, 1)
    write_rows(
        ws,
        2,
        [
            [
                p.get("full_name"),
                firm_name(p.get("firm_id")),
                p.get("role_title"),
                p.get("linkedin_url"),
                "",
                p.get("email"),
                "yes" if p.get("firm_id") in parts_by else "no",
            ]
            for p in li
        ],
    )

if "Generated" not in wb.sheetnames:
    g = wb.create_sheet("Generated")
    g["A1"] = "Generated from the shared book. Do not type in this file."
    g["A2"] = "The 17 Aug original was not modified."
    g["A3"] = datetime.now(timezone.utc).isoformat()
    g["A4"] = f"Firms {len(firms)} · people {len(people)} · participations {len(parts)}"

wb.save(out)
wb.close()
if ORIG.stat().st_mtime != stamp_before:
    raise SystemExit("ERROR: original workbook mtime changed")
print(json.dumps({
    "out": str(out),
    "firms": len(firms),
    "people": len(people),
    "participations": len(parts),
    "data_rows": len(sorted_firms),
    "orig_untouched": True,
}))
